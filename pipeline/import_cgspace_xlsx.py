"""CGSpace XLSX ingestion adapter for the EE Toolbox batch pipeline.

Reads yearly CGSpace XLSX files (produced by the CGSpace-extraction pipeline),
maps columns to the pipeline input schema, deduplicates by ``cgspace_id``, and
outputs a single JSON array file compatible with ``pipeline/run_batch.py --input``.

Usage::

    # Process all year files
    python -m pipeline.import_cgspace_xlsx \\
        --input-dir /path/to/xlsx/files \\
        --output items.json

    # Process specific years with a per-file row limit (testing)
    python -m pipeline.import_cgspace_xlsx \\
        --input-dir /path/to/xlsx/files \\
        --output items.json \\
        --years 2023,2024 \\
        --limit 50

    # Custom quality report path
    python -m pipeline.import_cgspace_xlsx \\
        --input-dir /path/to/xlsx/files \\
        --output items.json \\
        --report quality_report.txt
"""

import argparse
import json
import logging
import os
import re
import sys
from collections import Counter
from pathlib import Path
from typing import Optional

import openpyxl

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Column mapping: XLSX column name -> output JSON field name
# ---------------------------------------------------------------------------

COLUMN_MAP = {
    "dc.title": "title",
    "dcterms.abstract": "abstract",
    "dc.contributor.author": "authors",
    "dcterms.issued": "date",
    "dc.identifier.uri": "url",
    "dcterms.type": "doc_type",
    "item_handle": "cgspace_id",
    "cg.coverage.country": "_country",
    "cg.coverage.region": "_region",
}

# File naming pattern: cgspace_data_YYYY.xlsx
FILENAME_PATTERN = re.compile(r"^cgspace_data_(\d{4})\.xlsx$")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _safe_str(value) -> str:
    """Convert a cell value to a stripped string, treating None as empty."""
    if value is None:
        return ""
    return str(value).strip()


def _parse_semicolon_list(value) -> list[str]:
    """Split a semicolon-separated cell value into a deduplicated list."""
    raw = _safe_str(value)
    if not raw:
        return []
    parts = [p.strip() for p in raw.split(";") if p.strip()]
    # Preserve order while deduplicating
    seen = set()
    result = []
    for p in parts:
        if p not in seen:
            seen.add(p)
            result.append(p)
    return result


def _build_geography_hints(country_val, region_val) -> list[str]:
    """Combine country and region fields into a single geography_hints list."""
    countries = _parse_semicolon_list(country_val)
    regions = _parse_semicolon_list(region_val)
    # Countries first, then regions, deduplicated
    seen = set()
    hints = []
    for item in countries + regions:
        if item not in seen:
            seen.add(item)
            hints.append(item)
    return hints


def _discover_xlsx_files(
    input_dir: str, years: Optional[list[int]] = None
) -> list[tuple[int, str]]:
    """Find cgspace_data_YYYY.xlsx files, returning (year, filepath) pairs sorted by year."""
    results = []
    for name in os.listdir(input_dir):
        m = FILENAME_PATTERN.match(name)
        if m:
            year = int(m.group(1))
            if years is None or year in years:
                results.append((year, os.path.join(input_dir, name)))
    results.sort(key=lambda x: x[0])
    return results


def _read_xlsx_file(
    filepath: str, limit: Optional[int] = None
) -> tuple[list[dict], dict]:
    """Read a single XLSX file and return (items, column_index_map).

    Each item is a dict with the mapped output fields.  Rows without a title
    are skipped.  Within-file deduplication by cgspace_id keeps only the first
    occurrence (rows are duplicated when a document has multiple attachments).

    Args:
        filepath: Path to the XLSX file.
        limit: Maximum number of unique items to return (for testing).

    Returns:
        A tuple of (list_of_item_dicts, stats_dict).
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    # Map header names to column indices
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_indices = {}
    for xlsx_col, json_field in COLUMN_MAP.items():
        try:
            col_indices[json_field] = headers.index(xlsx_col)
        except ValueError:
            logger.warning(
                "Column '%s' not found in %s — field '%s' will be empty",
                xlsx_col, filepath, json_field,
            )
            col_indices[json_field] = None

    items = []
    seen_ids = set()
    stats = {
        "total_rows": 0,
        "skipped_no_title": 0,
        "skipped_duplicate": 0,
        "kept": 0,
    }

    for row in ws.iter_rows(min_row=2, values_only=True):
        stats["total_rows"] += 1

        # Extract title — skip rows without one
        title_idx = col_indices.get("title")
        title = _safe_str(row[title_idx]) if title_idx is not None else ""
        if not title:
            stats["skipped_no_title"] += 1
            continue

        # Extract cgspace_id for within-file dedup
        cid_idx = col_indices.get("cgspace_id")
        cgspace_id = _safe_str(row[cid_idx]) if cid_idx is not None else ""

        if cgspace_id and cgspace_id in seen_ids:
            stats["skipped_duplicate"] += 1
            continue
        if cgspace_id:
            seen_ids.add(cgspace_id)

        # Extract remaining fields
        def _get(field: str) -> str:
            idx = col_indices.get(field)
            if idx is None:
                return ""
            return _safe_str(row[idx])

        country_idx = col_indices.get("_country")
        region_idx = col_indices.get("_region")
        country_val = row[country_idx] if country_idx is not None else None
        region_val = row[region_idx] if region_idx is not None else None

        item = {
            "title": title,
            "abstract": _get("abstract"),
            "authors": _get("authors"),
            "date": _get("date"),
            "url": _get("url"),
            "doc_type": _get("doc_type"),
            "cgspace_id": cgspace_id,
            "geography_hints": _build_geography_hints(country_val, region_val),
        }

        items.append(item)
        stats["kept"] += 1

        if limit is not None and stats["kept"] >= limit:
            break

    wb.close()
    return items, stats


# ---------------------------------------------------------------------------
# Quality report
# ---------------------------------------------------------------------------


def _compute_quality_report(
    items: list[dict],
    per_year_stats: dict[int, dict],
    dedup_stats: dict,
) -> str:
    """Generate a data quality report as a formatted string."""
    lines = []
    lines.append("=" * 72)
    lines.append("CGSpace XLSX Import — Data Quality Report")
    lines.append("=" * 72)

    total = len(items)

    # --- 1. Summary ---
    lines.append("")
    lines.append("1. SUMMARY")
    lines.append(f"   Files processed:        {len(per_year_stats)}")
    lines.append(f"   Total records output:    {total}")

    if total > 0:
        with_title = sum(1 for i in items if i["title"])
        with_abstract = sum(1 for i in items if i["abstract"])
        lines.append(
            f"   Records with title:      {with_title} ({with_title / total:.1%})"
        )
        lines.append(
            f"   Records with abstract:   {with_abstract} ({with_abstract / total:.1%})"
        )

    # --- 2. Per-year breakdown ---
    lines.append("")
    lines.append("2. PER-YEAR BREAKDOWN")
    lines.append(
        f"   {'Year':<6} {'Rows':>7} {'Unique':>7} {'Dupes':>7} "
        f"{'Title%':>7} {'Abstr%':>7} {'Author%':>7} {'URL%':>7}"
    )
    lines.append("   " + "-" * 62)

    for year in sorted(per_year_stats.keys()):
        ys = per_year_stats[year]
        n = ys["kept"]
        if n == 0:
            lines.append(f"   {year:<6} {ys['total_rows']:>7} {0:>7} {0:>7}   (no records)")
            continue
        year_prefix = str(year)
        year_items = [i for i in items if i["date"].startswith(year_prefix)]
        n_items = len(year_items) if year_items else n
        pct_title = sum(1 for i in year_items if i["title"]) / n_items if n_items else 0
        pct_abstract = sum(1 for i in year_items if i["abstract"]) / n_items if n_items else 0
        pct_authors = sum(1 for i in year_items if i["authors"]) / n_items if n_items else 0
        pct_url = sum(1 for i in year_items if i["url"]) / n_items if n_items else 0
        lines.append(
            f"   {year:<6} {ys['total_rows']:>7} {ys['kept']:>7} "
            f"{ys['skipped_duplicate']:>7} "
            f"{pct_title:>6.1%} {pct_abstract:>6.1%} "
            f"{pct_authors:>6.1%} {pct_url:>6.1%}"
        )

    # --- 3. Field completeness ---
    lines.append("")
    lines.append("3. FIELD COMPLETENESS (across all output records)")
    if total > 0:
        fields = ["title", "abstract", "authors", "date", "url", "doc_type", "cgspace_id"]
        for f in fields:
            count = sum(1 for i in items if i[f])
            lines.append(f"   {f:<20} {count:>7} / {total:<7} ({count / total:.1%})")
        geo_count = sum(1 for i in items if i["geography_hints"])
        lines.append(
            f"   {'geography_hints':<20} {geo_count:>7} / {total:<7} ({geo_count / total:.1%})"
        )
    else:
        lines.append("   (no records)")

    # --- 4. Deduplication stats ---
    lines.append("")
    lines.append("4. DEDUPLICATION")
    lines.append(
        f"   Cross-file duplicates removed: {dedup_stats.get('cross_file_dupes', 0)}"
    )
    total_within = sum(s["skipped_duplicate"] for s in per_year_stats.values())
    lines.append(f"   Within-file duplicates removed: {total_within}")

    # --- 5. Top document types ---
    lines.append("")
    lines.append("5. TOP DOCUMENT TYPES")
    type_counts = Counter(i["doc_type"] for i in items if i["doc_type"])
    if type_counts:
        for dtype, count in type_counts.most_common(20):
            lines.append(f"   {dtype:<50} {count:>6}")
    else:
        lines.append("   (no document types found)")

    lines.append("")
    lines.append("=" * 72)
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Main logic
# ---------------------------------------------------------------------------


def import_cgspace_xlsx(
    input_dir: str,
    output_path: str,
    years: Optional[list[int]] = None,
    limit: Optional[int] = None,
    report_path: Optional[str] = None,
) -> list[dict]:
    """Read CGSpace XLSX files and write a pipeline-compatible JSON array.

    Args:
        input_dir: Directory containing ``cgspace_data_YYYY.xlsx`` files.
        output_path: Destination path for the JSON output.
        years: If provided, only process these years.
        limit: Max unique records per file (for testing).
        report_path: Path to write the quality report. Defaults to
                     ``<output>_quality_report.txt``.

    Returns:
        The list of item dicts written to the output file.
    """
    # Discover files
    files = _discover_xlsx_files(input_dir, years)
    if not files:
        logger.error(
            "No cgspace_data_YYYY.xlsx files found in %s (years filter: %s)",
            input_dir, years,
        )
        return []

    logger.info(
        "Found %d XLSX file(s) to process: years %s",
        len(files),
        [y for y, _ in files],
    )

    # Process files one at a time (memory-efficient)
    all_items: dict[str, dict] = {}  # keyed by cgspace_id for cross-file dedup
    per_year_stats: dict[int, dict] = {}
    cross_file_dupes = 0

    for year, filepath in files:
        logger.info("Processing %s (year %d)...", os.path.basename(filepath), year)
        items, stats = _read_xlsx_file(filepath, limit=limit)
        per_year_stats[year] = stats

        logger.info(
            "  Year %d: %d rows, %d unique items, %d within-file dupes, %d skipped (no title)",
            year, stats["total_rows"], stats["kept"],
            stats["skipped_duplicate"], stats["skipped_no_title"],
        )

        # Cross-file deduplication: keep latest (higher year wins)
        for item in items:
            cid = item["cgspace_id"]
            if not cid:
                # Items without a cgspace_id cannot be deduped — keep all
                # Use a synthetic key to avoid collisions
                cid = f"_notitle_{id(item)}"
                all_items[cid] = item
                continue

            if cid in all_items:
                existing_year = all_items[cid].get("date", "")
                new_year = item.get("date", "")
                if new_year >= existing_year:
                    all_items[cid] = item
                cross_file_dupes += 1
            else:
                all_items[cid] = item

    # Collect final items
    final_items = list(all_items.values())
    logger.info(
        "Total: %d unique items after deduplication (%d cross-file dupes removed)",
        len(final_items), cross_file_dupes,
    )

    # Ensure output directory exists
    output_dir = os.path.dirname(os.path.abspath(output_path))
    os.makedirs(output_dir, exist_ok=True)

    # Write JSON output
    with open(output_path, "w", encoding="utf-8") as fp:
        json.dump(final_items, fp, indent=2, ensure_ascii=False)
    logger.info("Wrote %d items to %s", len(final_items), output_path)

    # Generate quality report
    dedup_stats = {"cross_file_dupes": cross_file_dupes}
    report_text = _compute_quality_report(final_items, per_year_stats, dedup_stats)

    # Print to stdout
    print(report_text)

    # Write to file
    if report_path is None:
        base, ext = os.path.splitext(output_path)
        report_path = base + "_quality_report.txt"
    with open(report_path, "w", encoding="utf-8") as fp:
        fp.write(report_text + "\n")
    logger.info("Quality report written to %s", report_path)

    return final_items


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="python -m pipeline.import_cgspace_xlsx",
        description=(
            "Import CGSpace XLSX files into a JSON array compatible with "
            "the EE Toolbox batch pipeline (pipeline/run_batch.py --input)."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples:\n"
            "  # Process all year files:\n"
            "  python -m pipeline.import_cgspace_xlsx \\\n"
            "      --input-dir /path/to/xlsx --output items.json\n\n"
            "  # Process 2023-2024 with 50-row limit per file:\n"
            "  python -m pipeline.import_cgspace_xlsx \\\n"
            "      --input-dir /path/to/xlsx --output items.json \\\n"
            "      --years 2023,2024 --limit 50\n"
        ),
    )
    parser.add_argument(
        "--input-dir",
        required=True,
        help="Directory containing cgspace_data_YYYY.xlsx files.",
    )
    parser.add_argument(
        "--output",
        required=True,
        help="Output JSON file path.",
    )
    parser.add_argument(
        "--years",
        default=None,
        help="Comma-separated years to process (e.g., '2023,2024'). Default: all.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Max unique records per XLSX file (for testing).",
    )
    parser.add_argument(
        "--report",
        default=None,
        help="Path to write the data quality report. Default: <output>_quality_report.txt.",
    )
    return parser


def main() -> None:
    parser = _build_parser()
    args = parser.parse_args()

    # Configure logging
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    # Parse --years
    years = None
    if args.years:
        try:
            years = [int(y.strip()) for y in args.years.split(",")]
        except ValueError:
            print(
                f"Error: --years must be comma-separated integers, got: {args.years}",
                file=sys.stderr,
            )
            sys.exit(2)

    items = import_cgspace_xlsx(
        input_dir=args.input_dir,
        output_path=args.output,
        years=years,
        limit=args.limit,
        report_path=args.report,
    )

    if not items:
        logger.warning("No items produced. Check input directory and year filters.")
        sys.exit(1)

    logger.info("Done. %d items written.", len(items))


if __name__ == "__main__":
    main()
