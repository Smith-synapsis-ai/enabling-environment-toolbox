"""Admin Analytics router: 14 endpoints for the admin dashboard.

Endpoints:
    GET /admin/analytics/overview            -- KPI overview cards
    GET /admin/analytics/timeseries          -- time-series metrics
    GET /admin/analytics/top-tools           -- top tools by views/ratings/searches
    GET /admin/analytics/search-terms        -- popular search terms
    GET /admin/analytics/geography           -- geography breakdown
    GET /admin/analytics/engagement-funnel   -- engagement funnel stages
    GET /admin/analytics/sessions            -- session list with activity summary
    GET /admin/analytics/pathway-completion  -- actionable pathway funnel
    GET /admin/analytics/mau-growth          -- monthly active users growth
    GET /admin/analytics/pulse-survey-scores -- pulse survey average scores
    GET /admin/analytics/traffic-sources     -- UTM traffic source breakdown
    GET /admin/analytics/suggestion-uptake   -- search-to-view conversion rate
    GET /admin/analytics/broken-links        -- tools with potentially broken links
    GET /admin/analytics/export              -- XLSX export of analytics data
"""

import io
import logging
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.routers.admin import verify_admin_token

logger = logging.getLogger(__name__)

router = APIRouter()


# ---------------------------------------------------------------------------
# Helper: convert row values for JSON serialization
# ---------------------------------------------------------------------------

def _serialize_row(row: dict) -> dict:
    """Convert UUID and datetime values to strings for JSON serialization."""
    result = {}
    for key, value in row.items():
        if isinstance(value, datetime):
            result[key] = value.isoformat()
        elif hasattr(value, "hex") and callable(getattr(value, "hex", None)):
            # UUID-like objects
            result[key] = str(value)
        else:
            result[key] = value
    return result


# ---------------------------------------------------------------------------
# 1. GET /admin/analytics/overview
# ---------------------------------------------------------------------------

@router.get("/admin/analytics/overview")
async def analytics_overview(
    days: int = Query(30, ge=1, le=365),
    _user: str = Depends(verify_admin_token),
    db: AsyncSession = Depends(get_db),
):
    """Return KPI overview cards for the admin dashboard."""
    try:
        # Total non-bot sessions in period
        result = await db.execute(
            text("""
                SELECT COUNT(DISTINCT session_id) AS total_users
                FROM user_sessions
                WHERE started_at >= NOW() - (:days * INTERVAL '1 day')
                AND (is_bot = false OR is_bot IS NULL)
            """),
            {"days": days},
        )
        total_users = result.scalar() or 0

        # Active users: sessions with at least 1 search or view
        result = await db.execute(
            text("""
                SELECT COUNT(DISTINCT us.session_id) AS active_users
                FROM user_sessions us
                WHERE us.started_at >= NOW() - (:days * INTERVAL '1 day')
                AND (us.is_bot = false OR us.is_bot IS NULL)
                AND (
                    EXISTS (
                        SELECT 1 FROM search_logs sl
                        WHERE sl.session_id = us.session_id
                        AND sl.created_at >= NOW() - (:days * INTERVAL '1 day')
                    )
                    OR EXISTS (
                        SELECT 1 FROM tool_views tv
                        WHERE tv.session_id = us.session_id
                        AND tv.created_at >= NOW() - (:days * INTERVAL '1 day')
                    )
                )
            """),
            {"days": days},
        )
        active_users = result.scalar() or 0

        # Total searches
        result = await db.execute(
            text("""
                SELECT COUNT(*) FROM search_logs
                WHERE created_at >= NOW() - (:days * INTERVAL '1 day')
            """),
            {"days": days},
        )
        total_searches = result.scalar() or 0

        # Total views
        result = await db.execute(
            text("""
                SELECT COUNT(*) FROM tool_views
                WHERE created_at >= NOW() - (:days * INTERVAL '1 day')
            """),
            {"days": days},
        )
        total_views = result.scalar() or 0

        # Total ratings
        result = await db.execute(
            text("""
                SELECT COUNT(*) FROM rating_events
                WHERE created_at >= NOW() - (:days * INTERVAL '1 day')
            """),
            {"days": days},
        )
        total_ratings = result.scalar() or 0

        # Total email captures
        result = await db.execute(
            text("""
                SELECT COUNT(*) FROM email_captures
                WHERE created_at >= NOW() - (:days * INTERVAL '1 day')
            """),
            {"days": days},
        )
        total_emails = result.scalar() or 0

        # Total chat turns (user messages only)
        result = await db.execute(
            text("""
                SELECT COUNT(*) FROM conversation_turns
                WHERE created_at >= NOW() - (:days * INTERVAL '1 day')
                AND role = 'user'
            """),
            {"days": days},
        )
        total_chats = result.scalar() or 0

        # Total tools (all visible)
        result = await db.execute(
            text("SELECT COUNT(*) FROM tools WHERE is_visible = true")
        )
        total_tools = result.scalar() or 0

        return {
            "period_days": days,
            "total_users": total_users,
            "active_users": active_users,
            "total_searches": total_searches,
            "total_views": total_views,
            "total_ratings": total_ratings,
            "total_emails": total_emails,
            "total_chats": total_chats,
            "total_tools": total_tools,
        }
    except Exception as e:
        logger.exception("Error in analytics_overview")
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# 2. GET /admin/analytics/timeseries
# ---------------------------------------------------------------------------

@router.get("/admin/analytics/timeseries")
async def analytics_timeseries(
    days: int = Query(90, ge=1, le=365),
    granularity: str = Query("daily"),
    metric: str = Query("users"),
    _user: str = Depends(verify_admin_token),
    db: AsyncSession = Depends(get_db),
):
    """Return time-series data points for a given metric and granularity."""
    if granularity not in ("daily", "weekly", "monthly"):
        raise HTTPException(status_code=400, detail="granularity must be daily, weekly, or monthly")
    if metric not in ("users", "searches", "views", "ratings"):
        raise HTTPException(status_code=400, detail="metric must be users, searches, views, or ratings")

    # Map granularity to date_trunc argument
    trunc_map = {"daily": "day", "weekly": "week", "monthly": "month"}
    trunc_val = trunc_map[granularity]

    try:
        if metric == "users":
            result = await db.execute(
                text(f"""
                    SELECT date_trunc('{trunc_val}', started_at) AS date,
                           COUNT(DISTINCT session_id) AS value
                    FROM user_sessions
                    WHERE started_at >= NOW() - (:days * INTERVAL '1 day')
                    AND (is_bot = false OR is_bot IS NULL)
                    GROUP BY date
                    ORDER BY date
                """),
                {"days": days},
            )
        elif metric == "searches":
            result = await db.execute(
                text(f"""
                    SELECT date_trunc('{trunc_val}', created_at) AS date,
                           COUNT(*) AS value
                    FROM search_logs
                    WHERE created_at >= NOW() - (:days * INTERVAL '1 day')
                    GROUP BY date
                    ORDER BY date
                """),
                {"days": days},
            )
        elif metric == "views":
            result = await db.execute(
                text(f"""
                    SELECT date_trunc('{trunc_val}', created_at) AS date,
                           COUNT(*) AS value
                    FROM tool_views
                    WHERE created_at >= NOW() - (:days * INTERVAL '1 day')
                    GROUP BY date
                    ORDER BY date
                """),
                {"days": days},
            )
        else:  # ratings
            result = await db.execute(
                text(f"""
                    SELECT date_trunc('{trunc_val}', created_at) AS date,
                           COUNT(*) AS value
                    FROM rating_events
                    WHERE created_at >= NOW() - (:days * INTERVAL '1 day')
                    GROUP BY date
                    ORDER BY date
                """),
                {"days": days},
            )

        rows = result.mappings().all()
        data = [
            {"date": row["date"].isoformat() if row["date"] else None, "value": row["value"]}
            for row in rows
        ]
        return {"period_days": days, "granularity": granularity, "metric": metric, "data": data}
    except Exception as e:
        logger.exception("Error in analytics_timeseries")
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# 3. GET /admin/analytics/top-tools
# ---------------------------------------------------------------------------

@router.get("/admin/analytics/top-tools")
async def analytics_top_tools(
    days: int = Query(30, ge=1, le=365),
    limit: int = Query(10, ge=1, le=100),
    sort_by: str = Query("views"),
    _user: str = Depends(verify_admin_token),
    db: AsyncSession = Depends(get_db),
):
    """Return top tools ranked by views, ratings, or search appearances."""
    if sort_by not in ("views", "ratings", "searches"):
        raise HTTPException(status_code=400, detail="sort_by must be views, ratings, or searches")

    sort_column_map = {
        "views": "views",
        "ratings": "ratings_count",
        "searches": "searches",
    }
    sort_col = sort_column_map[sort_by]

    try:
        result = await db.execute(
            text(f"""
                SELECT
                    t.id AS tool_id,
                    t.title,
                    COALESCE(v.view_count, 0) AS views,
                    COALESCE(r.rating_count, 0) AS ratings_count,
                    COALESCE(r.avg_rating, 0) AS avg_rating,
                    COALESCE(s.search_count, 0) AS searches
                FROM tools t
                LEFT JOIN (
                    SELECT tool_id, COUNT(*) AS view_count
                    FROM tool_views
                    WHERE created_at >= NOW() - (:days * INTERVAL '1 day')
                    GROUP BY tool_id
                ) v ON v.tool_id = t.id
                LEFT JOIN (
                    SELECT tool_id, COUNT(*) AS rating_count, ROUND(AVG(rating), 2) AS avg_rating
                    FROM rating_events
                    WHERE created_at >= NOW() - (:days * INTERVAL '1 day')
                    GROUP BY tool_id
                ) r ON r.tool_id = t.id
                LEFT JOIN (
                    SELECT unnest(results_tool_ids) AS tool_id, COUNT(*) AS search_count
                    FROM search_logs
                    WHERE created_at >= NOW() - (:days * INTERVAL '1 day')
                    GROUP BY tool_id
                ) s ON s.tool_id = t.id
                WHERE t.is_visible = true
                ORDER BY {sort_col} DESC NULLS LAST
                LIMIT :limit
            """),
            {"days": days, "limit": limit},
        )
        rows = result.mappings().all()
        data = [
            {
                "tool_id": str(row["tool_id"]),
                "title": row["title"],
                "views": row["views"],
                "ratings_count": row["ratings_count"],
                "avg_rating": float(row["avg_rating"]),
                "searches": row["searches"],
            }
            for row in rows
        ]
        return {"period_days": days, "sort_by": sort_by, "tools": data}
    except Exception as e:
        logger.exception("Error in analytics_top_tools")
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# 4. GET /admin/analytics/search-terms
# ---------------------------------------------------------------------------

@router.get("/admin/analytics/search-terms")
async def analytics_search_terms(
    days: int = Query(30, ge=1, le=365),
    limit: int = Query(20, ge=1, le=100),
    _user: str = Depends(verify_admin_token),
    db: AsyncSession = Depends(get_db),
):
    """Return popular search terms grouped by normalized query text."""
    try:
        result = await db.execute(
            text("""
                SELECT
                    LOWER(TRIM(query_text)) AS query_text,
                    COUNT(*) AS count,
                    ROUND(AVG(COALESCE(results_count, 0)), 1) AS avg_results
                FROM search_logs
                WHERE created_at >= NOW() - (:days * INTERVAL '1 day')
                AND query_text IS NOT NULL
                AND TRIM(query_text) != ''
                GROUP BY LOWER(TRIM(query_text))
                ORDER BY count DESC
                LIMIT :limit
            """),
            {"days": days, "limit": limit},
        )
        rows = result.mappings().all()
        data = [
            {
                "query_text": row["query_text"],
                "count": row["count"],
                "avg_results": float(row["avg_results"]),
            }
            for row in rows
        ]
        return {"period_days": days, "terms": data}
    except Exception as e:
        logger.exception("Error in analytics_search_terms")
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# 5. GET /admin/analytics/geography
# ---------------------------------------------------------------------------

@router.get("/admin/analytics/geography")
async def analytics_geography(
    days: int = Query(30, ge=1, le=365),
    _user: str = Depends(verify_admin_token),
    db: AsyncSession = Depends(get_db),
):
    """Return geography breakdown: tool counts and view counts per geography."""
    try:
        result = await db.execute(
            text("""
                SELECT
                    geo,
                    COUNT(DISTINCT t.id) AS tool_count,
                    COALESCE(SUM(v.view_count), 0) AS view_count
                FROM tools t
                CROSS JOIN LATERAL unnest(t.geography) AS geo
                LEFT JOIN (
                    SELECT tool_id, COUNT(*) AS view_count
                    FROM tool_views
                    WHERE created_at >= NOW() - (:days * INTERVAL '1 day')
                    GROUP BY tool_id
                ) v ON v.tool_id = t.id
                WHERE t.is_visible = true
                GROUP BY geo
                ORDER BY view_count DESC
            """),
            {"days": days},
        )
        rows = result.mappings().all()
        data = [
            {
                "geography": row["geo"],
                "tool_count": row["tool_count"],
                "view_count": row["view_count"],
            }
            for row in rows
        ]
        return {"period_days": days, "geographies": data}
    except Exception as e:
        logger.exception("Error in analytics_geography")
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# 6. GET /admin/analytics/engagement-funnel
# ---------------------------------------------------------------------------

@router.get("/admin/analytics/engagement-funnel")
async def analytics_engagement_funnel(
    days: int = Query(30, ge=1, le=365),
    _user: str = Depends(verify_admin_token),
    db: AsyncSession = Depends(get_db),
):
    """Return engagement funnel stage counts."""
    try:
        interval_clause = "NOW() - (:days * INTERVAL '1 day')"
        params = {"days": days}

        # Sessions: all non-bot sessions
        result = await db.execute(
            text(f"""
                SELECT COUNT(DISTINCT session_id) AS cnt
                FROM user_sessions
                WHERE started_at >= {interval_clause}
                AND (is_bot = false OR is_bot IS NULL)
            """),
            params,
        )
        sessions_count = result.scalar() or 0

        # Searched: sessions with >= 1 search_log
        result = await db.execute(
            text(f"""
                SELECT COUNT(DISTINCT us.session_id) AS cnt
                FROM user_sessions us
                WHERE us.started_at >= {interval_clause}
                AND (us.is_bot = false OR us.is_bot IS NULL)
                AND EXISTS (
                    SELECT 1 FROM search_logs sl
                    WHERE sl.session_id = us.session_id
                    AND sl.created_at >= {interval_clause}
                )
            """),
            params,
        )
        searched_count = result.scalar() or 0

        # Viewed Tool: sessions with >= 1 tool_view
        result = await db.execute(
            text(f"""
                SELECT COUNT(DISTINCT us.session_id) AS cnt
                FROM user_sessions us
                WHERE us.started_at >= {interval_clause}
                AND (us.is_bot = false OR us.is_bot IS NULL)
                AND EXISTS (
                    SELECT 1 FROM tool_views tv
                    WHERE tv.session_id = us.session_id
                    AND tv.created_at >= {interval_clause}
                )
            """),
            params,
        )
        viewed_count = result.scalar() or 0

        # Rated: sessions with >= 1 rating_event
        result = await db.execute(
            text(f"""
                SELECT COUNT(DISTINCT us.session_id) AS cnt
                FROM user_sessions us
                WHERE us.started_at >= {interval_clause}
                AND (us.is_bot = false OR us.is_bot IS NULL)
                AND EXISTS (
                    SELECT 1 FROM rating_events re
                    WHERE re.session_id = us.session_id
                    AND re.created_at >= {interval_clause}
                )
            """),
            params,
        )
        rated_count = result.scalar() or 0

        # Email Captured: sessions with email_capture
        result = await db.execute(
            text(f"""
                SELECT COUNT(DISTINCT us.session_id) AS cnt
                FROM user_sessions us
                WHERE us.started_at >= {interval_clause}
                AND (us.is_bot = false OR us.is_bot IS NULL)
                AND EXISTS (
                    SELECT 1 FROM email_captures ec
                    WHERE ec.session_id = us.session_id
                    AND ec.created_at >= {interval_clause}
                )
            """),
            params,
        )
        email_count = result.scalar() or 0

        return {
            "period_days": days,
            "stages": [
                {"stage": "Sessions", "count": sessions_count},
                {"stage": "Searched", "count": searched_count},
                {"stage": "Viewed Tool", "count": viewed_count},
                {"stage": "Rated", "count": rated_count},
                {"stage": "Email Captured", "count": email_count},
            ],
        }
    except Exception as e:
        logger.exception("Error in analytics_engagement_funnel")
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# 7. GET /admin/analytics/sessions
# ---------------------------------------------------------------------------

@router.get("/admin/analytics/sessions")
async def analytics_sessions(
    days: int = Query(7, ge=1, le=365),
    limit: int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
    _user: str = Depends(verify_admin_token),
    db: AsyncSession = Depends(get_db),
):
    """Return session records with activity summary."""
    try:
        result = await db.execute(
            text("""
                SELECT
                    us.session_id,
                    us.started_at,
                    us.last_active_at,
                    us.user_email,
                    us.user_type,
                    us.is_bot,
                    us.user_agent,
                    COALESCE(sl.search_count, 0) AS search_count,
                    COALESCE(tv.view_count, 0) AS view_count,
                    COALESCE(re.rating_count, 0) AS rating_count
                FROM user_sessions us
                LEFT JOIN (
                    SELECT session_id, COUNT(*) AS search_count
                    FROM search_logs
                    WHERE created_at >= NOW() - (:days * INTERVAL '1 day')
                    GROUP BY session_id
                ) sl ON sl.session_id = us.session_id
                LEFT JOIN (
                    SELECT session_id, COUNT(*) AS view_count
                    FROM tool_views
                    WHERE created_at >= NOW() - (:days * INTERVAL '1 day')
                    GROUP BY session_id
                ) tv ON tv.session_id = us.session_id
                LEFT JOIN (
                    SELECT session_id, COUNT(*) AS rating_count
                    FROM rating_events
                    WHERE created_at >= NOW() - (:days * INTERVAL '1 day')
                    GROUP BY session_id
                ) re ON re.session_id = us.session_id
                WHERE us.started_at >= NOW() - (:days * INTERVAL '1 day')
                ORDER BY us.started_at DESC
                LIMIT :limit OFFSET :offset
            """),
            {"days": days, "limit": limit, "offset": offset},
        )
        rows = result.mappings().all()
        data = [
            {
                "session_id": row["session_id"],
                "started_at": row["started_at"].isoformat() if row["started_at"] else None,
                "last_active_at": row["last_active_at"].isoformat() if row["last_active_at"] else None,
                "user_email": row["user_email"],
                "user_type": row["user_type"],
                "is_bot": row["is_bot"],
                "user_agent": row["user_agent"],
                "search_count": row["search_count"],
                "view_count": row["view_count"],
                "rating_count": row["rating_count"],
            }
            for row in rows
        ]

        # Get total count for pagination
        count_result = await db.execute(
            text("""
                SELECT COUNT(*) FROM user_sessions
                WHERE started_at >= NOW() - (:days * INTERVAL '1 day')
            """),
            {"days": days},
        )
        total = count_result.scalar() or 0

        return {"period_days": days, "total": total, "limit": limit, "offset": offset, "sessions": data}
    except Exception as e:
        logger.exception("Error in analytics_sessions")
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# 8. GET /admin/analytics/pathway-completion
# ---------------------------------------------------------------------------

@router.get("/admin/analytics/pathway-completion")
async def analytics_pathway_completion(
    days: int = Query(30, ge=1, le=365),
    _user: str = Depends(verify_admin_token),
    db: AsyncSession = Depends(get_db),
):
    """Return actionable pathway funnel: Discovery -> Evaluation -> Application."""
    try:
        interval_clause = "NOW() - (:days * INTERVAL '1 day')"
        params = {"days": days}

        # Discovery: sessions that searched
        result = await db.execute(
            text(f"""
                SELECT COUNT(DISTINCT us.session_id) AS cnt
                FROM user_sessions us
                WHERE us.started_at >= {interval_clause}
                AND (us.is_bot = false OR us.is_bot IS NULL)
                AND EXISTS (
                    SELECT 1 FROM search_logs sl
                    WHERE sl.session_id = us.session_id
                    AND sl.created_at >= {interval_clause}
                )
            """),
            params,
        )
        discovery_count = result.scalar() or 0

        # Evaluation: sessions that searched AND then viewed a tool
        result = await db.execute(
            text(f"""
                SELECT COUNT(DISTINCT us.session_id) AS cnt
                FROM user_sessions us
                WHERE us.started_at >= {interval_clause}
                AND (us.is_bot = false OR us.is_bot IS NULL)
                AND EXISTS (
                    SELECT 1 FROM search_logs sl
                    WHERE sl.session_id = us.session_id
                    AND sl.created_at >= {interval_clause}
                )
                AND EXISTS (
                    SELECT 1 FROM tool_views tv
                    WHERE tv.session_id = us.session_id
                    AND tv.created_at >= {interval_clause}
                )
            """),
            params,
        )
        evaluation_count = result.scalar() or 0

        # Application: sessions that viewed AND (rated OR saved OR captured email)
        result = await db.execute(
            text(f"""
                SELECT COUNT(DISTINCT us.session_id) AS cnt
                FROM user_sessions us
                WHERE us.started_at >= {interval_clause}
                AND (us.is_bot = false OR us.is_bot IS NULL)
                AND EXISTS (
                    SELECT 1 FROM tool_views tv
                    WHERE tv.session_id = us.session_id
                    AND tv.created_at >= {interval_clause}
                )
                AND (
                    EXISTS (
                        SELECT 1 FROM rating_events re
                        WHERE re.session_id = us.session_id
                        AND re.created_at >= {interval_clause}
                    )
                    OR EXISTS (
                        SELECT 1 FROM tool_saves ts
                        WHERE ts.session_id = us.session_id
                        AND ts.created_at >= {interval_clause}
                    )
                    OR EXISTS (
                        SELECT 1 FROM email_captures ec
                        WHERE ec.session_id = us.session_id
                        AND ec.created_at >= {interval_clause}
                    )
                )
            """),
            params,
        )
        application_count = result.scalar() or 0

        return {
            "period_days": days,
            "stages": [
                {"stage": "Discovery", "count": discovery_count, "description": "Sessions that searched"},
                {"stage": "Evaluation", "count": evaluation_count, "description": "Sessions that searched and viewed a tool"},
                {"stage": "Application", "count": application_count, "description": "Sessions that viewed and took action (rated, saved, or email)"},
            ],
        }
    except Exception as e:
        logger.exception("Error in analytics_pathway_completion")
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# 9. GET /admin/analytics/mau-growth
# ---------------------------------------------------------------------------

@router.get("/admin/analytics/mau-growth")
async def analytics_mau_growth(
    _user: str = Depends(verify_admin_token),
    db: AsyncSession = Depends(get_db),
):
    """Return monthly active users with month-over-month growth percentage."""
    try:
        result = await db.execute(
            text("""
                WITH monthly AS (
                    SELECT
                        date_trunc('month', started_at) AS month,
                        COUNT(DISTINCT session_id) AS mau
                    FROM user_sessions
                    WHERE (is_bot = false OR is_bot IS NULL)
                    GROUP BY month
                    ORDER BY month
                )
                SELECT
                    month,
                    mau,
                    ROUND(
                        CASE
                            WHEN LAG(mau) OVER (ORDER BY month) IS NULL THEN NULL
                            WHEN LAG(mau) OVER (ORDER BY month) = 0 THEN NULL
                            ELSE ((mau - LAG(mau) OVER (ORDER BY month))::numeric
                                  / LAG(mau) OVER (ORDER BY month)) * 100
                        END, 1
                    ) AS mom_growth_pct
                FROM monthly
            """)
        )
        rows = result.mappings().all()
        data = [
            {
                "month": row["month"].isoformat() if row["month"] else None,
                "mau": row["mau"],
                "mom_growth_pct": float(row["mom_growth_pct"]) if row["mom_growth_pct"] is not None else None,
            }
            for row in rows
        ]
        return {"months": data}
    except Exception as e:
        logger.exception("Error in analytics_mau_growth")
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# 10. GET /admin/analytics/pulse-survey-scores
# ---------------------------------------------------------------------------

@router.get("/admin/analytics/pulse-survey-scores")
async def analytics_pulse_survey_scores(
    days: int = Query(30, ge=1, le=365),
    _user: str = Depends(verify_admin_token),
    db: AsyncSession = Depends(get_db),
):
    """Return average pulse survey scores per question key."""
    try:
        result = await db.execute(
            text("""
                SELECT
                    question_key,
                    ROUND(AVG(score), 2) AS avg_score,
                    COUNT(*) AS response_count
                FROM pulse_survey_responses
                WHERE created_at >= NOW() - (:days * INTERVAL '1 day')
                GROUP BY question_key
                ORDER BY question_key
            """),
            {"days": days},
        )
        rows = result.mappings().all()
        scores = [
            {
                "question_key": row["question_key"],
                "avg_score": float(row["avg_score"]),
                "response_count": row["response_count"],
            }
            for row in rows
        ]

        # Compute overall average
        if scores:
            total_weighted = sum(s["avg_score"] * s["response_count"] for s in scores)
            total_responses = sum(s["response_count"] for s in scores)
            overall_avg = round(total_weighted / total_responses, 2) if total_responses > 0 else 0.0
        else:
            overall_avg = 0.0

        return {
            "period_days": days,
            "scores": scores,
            "overall_avg": overall_avg,
        }
    except Exception as e:
        logger.exception("Error in analytics_pulse_survey_scores")
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# 11. GET /admin/analytics/traffic-sources
# ---------------------------------------------------------------------------

@router.get("/admin/analytics/traffic-sources")
async def analytics_traffic_sources(
    days: int = Query(30, ge=1, le=365),
    _user: str = Depends(verify_admin_token),
    db: AsyncSession = Depends(get_db),
):
    """Return UTM-based traffic source breakdown."""
    try:
        result = await db.execute(
            text("""
                SELECT
                    COALESCE(utm_source, 'direct') AS utm_source,
                    COALESCE(utm_medium, 'none') AS utm_medium,
                    COUNT(DISTINCT session_id) AS session_count
                FROM user_sessions
                WHERE started_at >= NOW() - (:days * INTERVAL '1 day')
                AND (is_bot = false OR is_bot IS NULL)
                GROUP BY COALESCE(utm_source, 'direct'), COALESCE(utm_medium, 'none')
                ORDER BY session_count DESC
            """),
            {"days": days},
        )
        rows = result.mappings().all()
        data = [
            {
                "utm_source": row["utm_source"],
                "utm_medium": row["utm_medium"],
                "session_count": row["session_count"],
            }
            for row in rows
        ]
        return {"period_days": days, "sources": data}
    except Exception as e:
        logger.exception("Error in analytics_traffic_sources")
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# 12. GET /admin/analytics/suggestion-uptake
# ---------------------------------------------------------------------------

@router.get("/admin/analytics/suggestion-uptake")
async def analytics_suggestion_uptake(
    days: int = Query(30, ge=1, le=365),
    _user: str = Depends(verify_admin_token),
    db: AsyncSession = Depends(get_db),
):
    """Return search-to-view conversion (suggestion uptake) rate."""
    try:
        interval_clause = "NOW() - (:days * INTERVAL '1 day')"
        params = {"days": days}

        # Sessions with at least one search
        result = await db.execute(
            text(f"""
                SELECT COUNT(DISTINCT us.session_id) AS cnt
                FROM user_sessions us
                WHERE us.started_at >= {interval_clause}
                AND (us.is_bot = false OR us.is_bot IS NULL)
                AND EXISTS (
                    SELECT 1 FROM search_logs sl
                    WHERE sl.session_id = us.session_id
                    AND sl.created_at >= {interval_clause}
                )
            """),
            params,
        )
        sessions_with_search = result.scalar() or 0

        # Sessions with search AND subsequent tool view
        result = await db.execute(
            text(f"""
                SELECT COUNT(DISTINCT us.session_id) AS cnt
                FROM user_sessions us
                WHERE us.started_at >= {interval_clause}
                AND (us.is_bot = false OR us.is_bot IS NULL)
                AND EXISTS (
                    SELECT 1 FROM search_logs sl
                    WHERE sl.session_id = us.session_id
                    AND sl.created_at >= {interval_clause}
                )
                AND EXISTS (
                    SELECT 1 FROM tool_views tv
                    WHERE tv.session_id = us.session_id
                    AND tv.created_at >= {interval_clause}
                )
            """),
            params,
        )
        sessions_with_search_then_view = result.scalar() or 0

        uptake_pct = round(
            (sessions_with_search_then_view / sessions_with_search * 100) if sessions_with_search > 0 else 0.0,
            1,
        )

        return {
            "period_days": days,
            "sessions_with_search": sessions_with_search,
            "sessions_with_search_then_view": sessions_with_search_then_view,
            "uptake_pct": uptake_pct,
        }
    except Exception as e:
        logger.exception("Error in analytics_suggestion_uptake")
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# 13. GET /admin/analytics/broken-links
# ---------------------------------------------------------------------------

@router.get("/admin/analytics/broken-links")
async def analytics_broken_links(
    _user: str = Depends(verify_admin_token),
    db: AsyncSession = Depends(get_db),
):
    """Return visible tools that have a source_url (potential broken link candidates).

    Note: This endpoint currently returns all visible tools with a source_url.
    Actual HTTP reachability checking is not yet implemented.
    """
    try:
        result = await db.execute(
            text("""
                SELECT id, title, source_url
                FROM tools
                WHERE is_visible = true
                AND source_url IS NOT NULL
                AND TRIM(source_url) != ''
                ORDER BY title
            """)
        )
        rows = result.mappings().all()
        data = [
            {
                "tool_id": str(row["id"]),
                "title": row["title"],
                "source_url": row["source_url"],
            }
            for row in rows
        ]
        return {
            "note": "This lists all visible tools with source URLs. HTTP reachability checking is not yet implemented.",
            "total": len(data),
            "tools": data,
        }
    except Exception as e:
        logger.exception("Error in analytics_broken_links")
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# 14. GET /admin/analytics/export
# ---------------------------------------------------------------------------

@router.get("/admin/analytics/export")
async def analytics_export(
    days: int = Query(30, ge=1, le=365),
    format: str = Query("xlsx"),
    _user: str = Depends(verify_admin_token),
    db: AsyncSession = Depends(get_db),
):
    """Export analytics data as an XLSX file."""
    if format != "xlsx":
        raise HTTPException(status_code=400, detail="Only xlsx format is currently supported")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()

        # --- Sheet 1: Overview ---
        ws_overview = wb.active
        ws_overview.title = "Overview"

        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")

        # Fetch overview data
        overview = await _fetch_overview_data(db, days)

        ws_overview.append(["Metric", "Value"])
        for cell in ws_overview[1]:
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for key, value in overview.items():
            if key != "period_days":
                label = key.replace("_", " ").title()
                ws_overview.append([label, value])

        ws_overview.column_dimensions["A"].width = 25
        ws_overview.column_dimensions["B"].width = 15

        # --- Sheet 2: Top Tools ---
        ws_tools = wb.create_sheet("Top Tools")
        tool_headers = ["Title", "Views", "Ratings Count", "Avg Rating", "Searches"]
        ws_tools.append(tool_headers)
        for cell in ws_tools[1]:
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        result = await db.execute(
            text("""
                SELECT
                    t.title,
                    COALESCE(v.view_count, 0) AS views,
                    COALESCE(r.rating_count, 0) AS ratings_count,
                    COALESCE(r.avg_rating, 0) AS avg_rating,
                    COALESCE(s.search_count, 0) AS searches
                FROM tools t
                LEFT JOIN (
                    SELECT tool_id, COUNT(*) AS view_count
                    FROM tool_views
                    WHERE created_at >= NOW() - (:days * INTERVAL '1 day')
                    GROUP BY tool_id
                ) v ON v.tool_id = t.id
                LEFT JOIN (
                    SELECT tool_id, COUNT(*) AS rating_count, ROUND(AVG(rating), 2) AS avg_rating
                    FROM rating_events
                    WHERE created_at >= NOW() - (:days * INTERVAL '1 day')
                    GROUP BY tool_id
                ) r ON r.tool_id = t.id
                LEFT JOIN (
                    SELECT unnest(results_tool_ids) AS tool_id, COUNT(*) AS search_count
                    FROM search_logs
                    WHERE created_at >= NOW() - (:days * INTERVAL '1 day')
                    GROUP BY tool_id
                ) s ON s.tool_id = t.id
                WHERE t.is_visible = true
                ORDER BY views DESC
                LIMIT 50
            """),
            {"days": days},
        )
        for row in result.mappings().all():
            ws_tools.append([
                row["title"],
                row["views"],
                row["ratings_count"],
                float(row["avg_rating"]),
                row["searches"],
            ])

        ws_tools.column_dimensions["A"].width = 40
        for col in ["B", "C", "D", "E"]:
            ws_tools.column_dimensions[col].width = 15

        # --- Sheet 3: Search Terms ---
        ws_search = wb.create_sheet("Search Terms")
        search_headers = ["Query Text", "Count", "Avg Results"]
        ws_search.append(search_headers)
        for cell in ws_search[1]:
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        result = await db.execute(
            text("""
                SELECT
                    LOWER(TRIM(query_text)) AS query_text,
                    COUNT(*) AS count,
                    ROUND(AVG(COALESCE(results_count, 0)), 1) AS avg_results
                FROM search_logs
                WHERE created_at >= NOW() - (:days * INTERVAL '1 day')
                AND query_text IS NOT NULL
                AND TRIM(query_text) != ''
                GROUP BY LOWER(TRIM(query_text))
                ORDER BY count DESC
                LIMIT 50
            """),
            {"days": days},
        )
        for row in result.mappings().all():
            ws_search.append([
                row["query_text"],
                row["count"],
                float(row["avg_results"]),
            ])

        ws_search.column_dimensions["A"].width = 40
        ws_search.column_dimensions["B"].width = 12
        ws_search.column_dimensions["C"].width = 15

        # --- Sheet 4: Sessions ---
        ws_sessions = wb.create_sheet("Sessions")
        session_headers = [
            "Session ID", "Started At", "Last Active At", "User Email",
            "User Type", "Is Bot", "Search Count", "View Count", "Rating Count",
        ]
        ws_sessions.append(session_headers)
        for cell in ws_sessions[1]:
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        result = await db.execute(
            text("""
                SELECT
                    us.session_id,
                    us.started_at,
                    us.last_active_at,
                    us.user_email,
                    us.user_type,
                    us.is_bot,
                    COALESCE(sl.search_count, 0) AS search_count,
                    COALESCE(tv.view_count, 0) AS view_count,
                    COALESCE(re.rating_count, 0) AS rating_count
                FROM user_sessions us
                LEFT JOIN (
                    SELECT session_id, COUNT(*) AS search_count
                    FROM search_logs
                    WHERE created_at >= NOW() - (:days * INTERVAL '1 day')
                    GROUP BY session_id
                ) sl ON sl.session_id = us.session_id
                LEFT JOIN (
                    SELECT session_id, COUNT(*) AS view_count
                    FROM tool_views
                    WHERE created_at >= NOW() - (:days * INTERVAL '1 day')
                    GROUP BY session_id
                ) tv ON tv.session_id = us.session_id
                LEFT JOIN (
                    SELECT session_id, COUNT(*) AS rating_count
                    FROM rating_events
                    WHERE created_at >= NOW() - (:days * INTERVAL '1 day')
                    GROUP BY session_id
                ) re ON re.session_id = us.session_id
                WHERE us.started_at >= NOW() - (:days * INTERVAL '1 day')
                ORDER BY us.started_at DESC
                LIMIT 200
            """),
            {"days": days},
        )
        for row in result.mappings().all():
            ws_sessions.append([
                row["session_id"],
                row["started_at"].isoformat() if row["started_at"] else "",
                row["last_active_at"].isoformat() if row["last_active_at"] else "",
                row["user_email"] or "",
                row["user_type"] or "",
                "Yes" if row["is_bot"] else "No",
                row["search_count"],
                row["view_count"],
                row["rating_count"],
            ])

        ws_sessions.column_dimensions["A"].width = 40
        ws_sessions.column_dimensions["B"].width = 22
        ws_sessions.column_dimensions["C"].width = 22
        ws_sessions.column_dimensions["D"].width = 30
        ws_sessions.column_dimensions["E"].width = 12
        ws_sessions.column_dimensions["F"].width = 8

        # Write to buffer and return
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=analytics-export-{days}d.xlsx"},
        )
    except Exception as e:
        logger.exception("Error in analytics_export")
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# Internal helper for export (reuses overview logic)
# ---------------------------------------------------------------------------

async def _fetch_overview_data(db: AsyncSession, days: int) -> dict:
    """Fetch overview KPI data for use in the export sheet."""
    params = {"days": days}

    result = await db.execute(
        text("""
            SELECT COUNT(DISTINCT session_id) AS total_users
            FROM user_sessions
            WHERE started_at >= NOW() - (:days * INTERVAL '1 day')
            AND (is_bot = false OR is_bot IS NULL)
        """),
        params,
    )
    total_users = result.scalar() or 0

    result = await db.execute(
        text("""
            SELECT COUNT(DISTINCT us.session_id) AS active_users
            FROM user_sessions us
            WHERE us.started_at >= NOW() - (:days * INTERVAL '1 day')
            AND (us.is_bot = false OR us.is_bot IS NULL)
            AND (
                EXISTS (SELECT 1 FROM search_logs sl WHERE sl.session_id = us.session_id AND sl.created_at >= NOW() - (:days * INTERVAL '1 day'))
                OR EXISTS (SELECT 1 FROM tool_views tv WHERE tv.session_id = us.session_id AND tv.created_at >= NOW() - (:days * INTERVAL '1 day'))
            )
        """),
        params,
    )
    active_users = result.scalar() or 0

    result = await db.execute(
        text("SELECT COUNT(*) FROM search_logs WHERE created_at >= NOW() - (:days * INTERVAL '1 day')"),
        params,
    )
    total_searches = result.scalar() or 0

    result = await db.execute(
        text("SELECT COUNT(*) FROM tool_views WHERE created_at >= NOW() - (:days * INTERVAL '1 day')"),
        params,
    )
    total_views = result.scalar() or 0

    result = await db.execute(
        text("SELECT COUNT(*) FROM rating_events WHERE created_at >= NOW() - (:days * INTERVAL '1 day')"),
        params,
    )
    total_ratings = result.scalar() or 0

    result = await db.execute(
        text("SELECT COUNT(*) FROM email_captures WHERE created_at >= NOW() - (:days * INTERVAL '1 day')"),
        params,
    )
    total_emails = result.scalar() or 0

    result = await db.execute(
        text("SELECT COUNT(*) FROM conversation_turns WHERE created_at >= NOW() - (:days * INTERVAL '1 day') AND role = 'user'"),
        params,
    )
    total_chats = result.scalar() or 0

    result = await db.execute(
        text("SELECT COUNT(*) FROM tools WHERE is_visible = true")
    )
    total_tools = result.scalar() or 0

    return {
        "period_days": days,
        "total_users": total_users,
        "active_users": active_users,
        "total_searches": total_searches,
        "total_views": total_views,
        "total_ratings": total_ratings,
        "total_emails": total_emails,
        "total_chats": total_chats,
        "total_tools": total_tools,
    }
